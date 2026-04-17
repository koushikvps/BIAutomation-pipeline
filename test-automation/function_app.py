"""
Test Automation Function App: AI-powered test framework for Power Apps + Data validation.

Separated from the BI Pipeline app as an independent product.

Endpoints:
  POST /api/run-tests         -> starts test orchestration
  GET  /api/test-progress     -> step-by-step progress
  POST /api/approve-test-plan -> approve pending test plan
  POST /api/decline-test-plan -> decline pending test plan
  POST /api/ado-webhook       -> auto-trigger on ADO state change
  GET  /api/download-tests    -> download Playwright test ZIP
  POST /api/upload-results    -> upload JUnit XML results
  POST /api/agent-log         -> receive log from local agent
  GET  /api/agent-poll        -> local agent job polling
  POST /api/queue-agent-job   -> queue UI test job
  POST /api/agent-results     -> upload local agent results
  GET  /api/download-data-report -> Excel report for data tests
  CRUD /api/test-categories   -> custom data test category management (SQL-backed)
  GET  /api/health            -> health check
"""

from __future__ import annotations

import json
import logging
import os
from datetime import datetime, timedelta, timezone

import azure.functions as func
import azure.durable_functions as df

from shared.config import AppConfig

app = df.DFApp(http_auth_level=func.AuthLevel.FUNCTION)
logger = logging.getLogger(__name__)

# In-memory stores
# WARNING: These in-memory stores are volatile and will be lost on Function App
# restart or scale-out. Data is not shared across instances.
# TODO: Replace with persistent storage (e.g., Azure Table Storage, Redis Cache,
# or Cosmos DB) for production use.
_agent_logs: dict[str, list[dict]] = {}
_pending_agent_jobs: list[dict] = []
_DEFAULT_TEST_CATEGORIES: list[dict] = [
    {"id": "completeness", "name": "Completeness", "description": "Verify all expected records exist", "enabled": True, "is_default": True},
    {"id": "accuracy", "name": "Accuracy", "description": "Verify data values match source", "enabled": True, "is_default": True},
    {"id": "timeliness", "name": "Timeliness", "description": "Verify data freshness within SLA", "enabled": True, "is_default": True},
    {"id": "uniqueness", "name": "Uniqueness", "description": "No duplicate records on key columns", "enabled": True, "is_default": True},
    {"id": "consistency", "name": "Consistency", "description": "Cross-layer alignment checks", "enabled": True, "is_default": True},
    {"id": "validity", "name": "Validity", "description": "Values conform to business rules", "enabled": True, "is_default": True},
    {"id": "referential_integrity", "name": "Referential Integrity", "description": "Foreign keys resolve", "enabled": True, "is_default": True},
]


# ============================================================
# SQL HELPERS: Test Categories Persistence
# ============================================================
def _get_config_db_connection():
    """Get pyodbc connection for the config database."""
    import pyodbc
    conn_str = os.environ.get("SQL_CONNECTION_STRING")
    if not conn_str:
        server = os.environ.get("CONFIG_DB_SERVER", os.environ.get("SOURCE_DB_SERVER", ""))
        db = os.environ.get("CONFIG_DB_NAME", os.environ.get("SOURCE_DB_NAME", ""))
        user = os.environ.get("SQL_ADMIN_USER", "sqladmin")
        pwd = os.environ.get("SQL_ADMIN_PASSWORD", "")
        conn_str = (
            f"DRIVER={{ODBC Driver 18 for SQL Server}};"
            f"SERVER=tcp:{server},1433;"
            f"DATABASE={db};"
            f"UID={user};PWD={pwd};"
            f"Encrypt=yes;TrustServerCertificate=no;Connection Timeout=10;"
        )
    return pyodbc.connect(conn_str, timeout=10)


def _ensure_categories_table():
    """Create test_categories table if it does not exist and seed defaults."""
    conn = _get_config_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            IF NOT EXISTS (
                SELECT * FROM INFORMATION_SCHEMA.TABLES
                WHERE TABLE_NAME = 'test_categories'
            )
            BEGIN
                CREATE TABLE test_categories (
                    id VARCHAR(100) PRIMARY KEY,
                    name VARCHAR(255) NOT NULL,
                    description VARCHAR(1000) DEFAULT '',
                    enabled BIT DEFAULT 1,
                    is_default BIT DEFAULT 0,
                    created_at DATETIME DEFAULT GETUTCDATE()
                )
            END
        """)
        conn.commit()
        # Seed defaults if table is empty
        cursor.execute("SELECT COUNT(*) FROM test_categories")
        count = cursor.fetchone()[0]
        if count == 0:
            for cat in _DEFAULT_TEST_CATEGORIES:
                cursor.execute(
                    "INSERT INTO test_categories (id, name, description, enabled, is_default) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (cat["id"], cat["name"], cat["description"],
                     cat["enabled"], cat.get("is_default", True)),
                )
            conn.commit()
    finally:
        conn.close()


def _get_categories_from_db() -> list[dict]:
    """Fetch all test categories from SQL, creating table and seeding if needed."""
    try:
        _ensure_categories_table()
        conn = _get_config_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT id, name, description, enabled, is_default, created_at "
                "FROM test_categories"
            )
            rows = cursor.fetchall()
            return [
                {
                    "id": r[0], "name": r[1], "description": r[2] or "",
                    "enabled": bool(r[3]), "is_default": bool(r[4]),
                    "created_at": r[5].isoformat() if r[5] else None,
                }
                for r in rows
            ]
        finally:
            conn.close()
    except Exception as e:
        logger.warning("Failed to read categories from SQL, using defaults: %s", e)
        return list(_DEFAULT_TEST_CATEGORIES)


# ============================================================
# HEALTH CHECK
# ============================================================
@app.route(route="health", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
async def health_check(req: func.HttpRequest) -> func.HttpResponse:
    checks = {"timestamp": datetime.now(timezone.utc).isoformat(), "status": "healthy", "product": "test-automation", "checks": {}}

    try:
        config = AppConfig.from_env()
        from shared.llm_client import LLMClient
        llm = LLMClient(config)
        llm.chat("You are a test.", "Reply with OK", max_tokens=5)
        checks["checks"]["llm"] = {"status": "ok"}
    except Exception as e:
        checks["checks"]["llm"] = {"status": "error", "message": str(e)[:200]}
        checks["status"] = "degraded"

    status_code = 200 if checks["status"] == "healthy" else 207
    return func.HttpResponse(json.dumps(checks, indent=2), mimetype="application/json", status_code=status_code)


# ============================================================
# HTTP: Start Test Pipeline
# ============================================================
@app.route(route="run-tests", methods=["POST"])
@app.durable_client_input(client_name="client")
async def run_tests(req: func.HttpRequest, client) -> func.HttpResponse:
    try:
        body = req.get_json()
        work_item_id = body.get("work_item_id")
        app_url = body.get("app_url", os.environ.get("POWERAPP_URL", ""))

        if not work_item_id:
            return func.HttpResponse('{"error": "work_item_id required"}', status_code=400, mimetype="application/json")

        from shared.ado_client import ADOClient
        ado = ADOClient()
        wi_fields = ado.get_work_item_fields(int(work_item_id))

        test_type = body.get("test_type", "")  # "ui" or "data"; skip AI routing
        custom_categories = body.get("custom_categories", [])

        test_input = {
            "work_item_id": int(work_item_id),
            "story_id": f"STORY-{work_item_id}",
            "title": wi_fields.get("title", ""),
            "description": wi_fields.get("description", ""),
            "acceptance_criteria": wi_fields.get("acceptance_criteria", ""),
            "app_url": app_url,
            "test_type": test_type,
            "custom_categories": custom_categories,
        }

        instance_id = await client.start_new("test_orchestrator", client_input=test_input)
        logger.info("Started test orchestrator %s for WI-%s", instance_id, work_item_id)

        return func.HttpResponse(json.dumps({
            "instance_id": instance_id,
            "story_id": test_input["story_id"],
            "work_item_id": str(work_item_id),
            "title": test_input["title"],
            "status": "STARTED",
        }), mimetype="application/json")

    except Exception as e:
        logger.error("run-tests error: %s", e)
        return func.HttpResponse(json.dumps({"error": str(e)}), status_code=500, mimetype="application/json")


# ============================================================
# HTTP: Test Progress
# ============================================================
@app.route(route="test-progress", methods=["GET"])
@app.durable_client_input(client_name="client")
async def test_progress(req: func.HttpRequest, client) -> func.HttpResponse:
    instance_id = req.params.get("instance_id", "")
    if not instance_id:
        return func.HttpResponse('{"error": "instance_id required"}', status_code=400, mimetype="application/json")
    try:
        status = await client.get_status(instance_id)
        if not status:
            return func.HttpResponse('{"error": "Instance not found"}', status_code=404, mimetype="application/json")
        runtime = status.runtime_status.name if status.runtime_status else "Unknown"
        custom = status.custom_status or {}
        is_complete = runtime in ("Completed", "Failed", "Terminated")
        return func.HttpResponse(json.dumps({
            "instance_id": instance_id,
            "runtime_status": runtime,
            "is_complete": is_complete,
            "steps": custom.get("steps", []),
            "story_id": custom.get("story_id", ""),
            "results": status.output if is_complete else None,
            "test_cases": custom.get("test_cases", []),
            "awaiting_test_review": custom.get("awaiting_test_review", False),
        }, default=str), mimetype="application/json")
    except Exception as e:
        return func.HttpResponse(json.dumps({"error": str(e)}), status_code=500, mimetype="application/json")


# ============================================================
# HTTP: ADO Webhook (auto-trigger on state change)
# ============================================================
@app.route(route="ado-webhook", methods=["POST"])
@app.durable_client_input(client_name="client")
async def ado_webhook(req: func.HttpRequest, client) -> func.HttpResponse:
    try:
        body = req.get_json()
        resource = body.get("resource", {})
        fields = resource.get("fields", {})
        new_state = ""
        if "System.State" in fields:
            new_state = fields["System.State"].get("newValue", "")
        if new_state.lower() not in ("ready for testing", "ready for test", "testing"):
            return func.HttpResponse(json.dumps({"status": "ignored", "reason": f"State '{new_state}' not a test trigger"}), mimetype="application/json")

        work_item_id = resource.get("workItemId") or resource.get("id")
        if not work_item_id:
            work_item_id = resource.get("revision", {}).get("id")
        if not work_item_id:
            return func.HttpResponse('{"error": "Could not extract work item ID"}', status_code=400, mimetype="application/json")

        from shared.ado_client import ADOClient
        ado = ADOClient()
        wi_fields = ado.get_work_item_fields(int(work_item_id))

        test_input = {
            "work_item_id": int(work_item_id),
            "story_id": f"STORY-{work_item_id}",
            "title": wi_fields.get("title", ""),
            "description": wi_fields.get("description", ""),
            "acceptance_criteria": wi_fields.get("acceptance_criteria", ""),
            "app_url": os.environ.get("POWERAPP_URL", ""),
            "auto_triggered": True,
        }
        instance_id = await client.start_new("test_orchestrator", client_input=test_input)
        logger.info("Auto-triggered test pipeline for WI-%s", work_item_id)
        return func.HttpResponse(json.dumps({"status": "triggered", "instance_id": instance_id, "work_item_id": str(work_item_id)}), mimetype="application/json")
    except Exception as e:
        logger.error("ADO webhook error: %s", e)
        return func.HttpResponse(json.dumps({"error": str(e)}), status_code=500, mimetype="application/json")


# ============================================================
# HTTP: Download Tests ZIP
# ============================================================
@app.route(route="download-tests", methods=["GET"])
@app.durable_client_input(client_name="client")
async def download_tests(req: func.HttpRequest, client) -> func.HttpResponse:
    import io, zipfile
    instance_id = req.params.get("instance_id", "")
    if not instance_id:
        return func.HttpResponse('{"error": "instance_id required"}', status_code=400, mimetype="application/json")
    try:
        status = await client.get_status(instance_id)
        if not status:
            return func.HttpResponse('{"error": "Instance not found"}', status_code=404, mimetype="application/json")
        custom = status.custom_status or {}
        generated_code = custom.get("generated_code")
        app_url = custom.get("app_url", "")
        story_id = custom.get("story_id", "")
        if not generated_code:
            return func.HttpResponse('{"error": "Tests not generated yet"}', status_code=400, mimetype="application/json")

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            if generated_code.get("conftest"):
                zf.writestr("conftest.py", generated_code["conftest"])
            for tf in generated_code.get("test_files", []):
                zf.writestr(tf["filename"], tf["content"])
            for po in generated_code.get("page_objects", []):
                zf.writestr(po["filename"], po["content"])
            zf.writestr("pytest.ini", "[pytest]\nasyncio_mode = auto\n")
            zf.writestr("run_tests.bat", f"@echo off\npython -m pytest --tb=short --junitxml=results.xml --browser-channel msedge -v\npause\n")
            zf.writestr("setup_once.bat", "@echo off\npip install pytest playwright pytest-playwright pytest-asyncio\nplaywright install msedge\npause\n")
        buf.seek(0)
        return func.HttpResponse(buf.read(), mimetype="application/zip", headers={"Content-Disposition": f'attachment; filename="tests_{story_id}.zip"'})
    except Exception as e:
        logger.error("download-tests error: %s", e)
        return func.HttpResponse(json.dumps({"error": str(e)}), status_code=500, mimetype="application/json")


# ============================================================
# HTTP: Upload JUnit XML Results
# ============================================================
@app.route(route="upload-results", methods=["POST"])
@app.durable_client_input(client_name="client")
async def upload_results(req: func.HttpRequest, client) -> func.HttpResponse:
    instance_id = req.params.get("instance_id", "")
    if not instance_id:
        return func.HttpResponse('{"error": "instance_id required"}', status_code=400, mimetype="application/json")
    try:
        import xml.etree.ElementTree as ET
        body = req.get_body().decode("utf-8", errors="replace")
        root = ET.fromstring(body)
        suite = root if root.tag == "testsuite" else root.find("testsuite")
        if suite is None:
            return func.HttpResponse('{"error": "Invalid JUnit XML"}', status_code=400, mimetype="application/json")

        total = int(suite.get("tests", 0))
        failures = int(suite.get("failures", 0))
        errors = int(suite.get("errors", 0))
        passed = total - failures - errors
        test_results = []
        for tc in suite.findall("testcase"):
            status_val = "passed"
            message = ""
            failure = tc.find("failure")
            error = tc.find("error")
            if failure is not None:
                status_val = "failed"
                message = failure.get("message", "")[:500]
            elif error is not None:
                status_val = "error"
                message = error.get("message", "")[:500]
            test_results.append({"name": tc.get("name", ""), "status": status_val, "duration": float(tc.get("time", 0)), "message": message})

        exec_result = {"status": "completed", "passed": passed, "failed": failures, "errors": errors, "total": total, "test_results": test_results}
        await client.raise_event(instance_id, "TestResultsUploaded", exec_result)
        return func.HttpResponse(json.dumps({"status": "uploaded", "instance_id": instance_id, "passed": passed, "failed": failures, "errors": errors, "total": total}), mimetype="application/json")
    except Exception as e:
        logger.error("upload-results error: %s", e)
        return func.HttpResponse(json.dumps({"error": str(e)}), status_code=500, mimetype="application/json")


# ============================================================
# HTTP: Approve / Decline Test Plan
# ============================================================
@app.route(route="approve-test-plan", methods=["POST"])
@app.durable_client_input(client_name="client")
async def approve_test_plan(req: func.HttpRequest, client) -> func.HttpResponse:
    try:
        body = req.get_json()
        instance_id = body.get("instance_id", "")
        if not instance_id:
            return func.HttpResponse(
                '{"error": "instance_id required"}',
                status_code=400, mimetype="application/json",
            )
        await client.raise_event(instance_id, "TestPlanApproved", {"approved": True})
        logger.info("Test plan approved for instance %s", instance_id)
        return func.HttpResponse(
            json.dumps({"status": "approved", "instance_id": instance_id}),
            mimetype="application/json",
        )
    except Exception as e:
        logger.error("approve-test-plan error: %s", e)
        return func.HttpResponse(json.dumps({"error": str(e)}), status_code=500, mimetype="application/json")


@app.route(route="decline-test-plan", methods=["POST"])
@app.durable_client_input(client_name="client")
async def decline_test_plan(req: func.HttpRequest, client) -> func.HttpResponse:
    try:
        body = req.get_json()
        instance_id = body.get("instance_id", "")
        if not instance_id:
            return func.HttpResponse(
                '{"error": "instance_id required"}',
                status_code=400, mimetype="application/json",
            )
        reason = body.get("reason", "")
        await client.raise_event(
            instance_id, "TestPlanApproved", {"approved": False, "reason": reason},
        )
        logger.info("Test plan declined for instance %s: %s", instance_id, reason)
        return func.HttpResponse(
            json.dumps({"status": "declined", "instance_id": instance_id, "reason": reason}),
            mimetype="application/json",
        )
    except Exception as e:
        logger.error("decline-test-plan error: %s", e)
        return func.HttpResponse(json.dumps({"error": str(e)}), status_code=500, mimetype="application/json")


# ============================================================
# LOCAL AGENT: Polling + Live Log + Results
# ============================================================
@app.route(route="agent-log", methods=["POST", "GET"], auth_level=func.AuthLevel.FUNCTION)
async def agent_log(req: func.HttpRequest) -> func.HttpResponse:
    if req.method == "GET":
        instance_id = req.params.get("instance_id", "")
        since = int(req.params.get("since", "0"))
        logs = _agent_logs.get(instance_id, [])
        return func.HttpResponse(json.dumps({"logs": [l for l in logs if l.get("idx", 0) > since]}), mimetype="application/json")
    body = req.get_json()
    instance_id = body.get("instance_id", "")
    if instance_id not in _agent_logs:
        _agent_logs[instance_id] = []
    idx = len(_agent_logs[instance_id]) + 1
    _agent_logs[instance_id].append({"idx": idx, "ts": datetime.now(timezone.utc).isoformat(), "level": body.get("level", "info"), "message": body.get("message", "")})
    if len(_agent_logs[instance_id]) > 500:
        _agent_logs[instance_id] = _agent_logs[instance_id][-500:]
    return func.HttpResponse(json.dumps({"ok": True}), mimetype="application/json")


@app.route(route="agent-poll", methods=["GET"], auth_level=func.AuthLevel.FUNCTION)
async def agent_poll(req: func.HttpRequest) -> func.HttpResponse:
    if _pending_agent_jobs:
        return func.HttpResponse(json.dumps({"job": _pending_agent_jobs.pop(0)}), mimetype="application/json")
    return func.HttpResponse(json.dumps({"job": None}), mimetype="application/json")


@app.route(route="queue-agent-job", methods=["POST"], auth_level=func.AuthLevel.FUNCTION)
async def queue_agent_job(req: func.HttpRequest) -> func.HttpResponse:
    body = req.get_json()
    instance_id = body.get("instance_id", "")
    if not instance_id:
        return func.HttpResponse(json.dumps({"error": "instance_id required"}), status_code=400, mimetype="application/json")
    if not any(j.get("instance_id") == instance_id for j in _pending_agent_jobs):
        _pending_agent_jobs.append({"instance_id": instance_id, "story_id": body.get("story_id", "")})
    return func.HttpResponse(json.dumps({"queued": True}), mimetype="application/json")


@app.route(route="agent-results", methods=["POST"], auth_level=func.AuthLevel.FUNCTION)
@app.durable_client_input(client_name="client")
async def agent_results(req: func.HttpRequest, client) -> func.HttpResponse:
    try:
        instance_id = req.params.get("instance_id", "")
        body = req.get_json()
        await client.raise_event(instance_id, "TestResultsUploaded", body)
        return func.HttpResponse(json.dumps({"status": "received"}), mimetype="application/json")
    except Exception as e:
        logger.error("agent-results error: %s", e)
        return func.HttpResponse(json.dumps({"error": str(e)}), status_code=500, mimetype="application/json")


# ============================================================
# HTTP: Download Data Test Report (Excel)
# ============================================================
@app.route(route="download-data-report", methods=["GET"], auth_level=func.AuthLevel.FUNCTION)
@app.durable_client_input(client_name="client")
async def download_data_report(req: func.HttpRequest, client) -> func.HttpResponse:
    try:
        instance_id = req.params.get("instance_id", "")
        if not instance_id:
            return func.HttpResponse("instance_id required", status_code=400)
        status = await client.get_status(instance_id)
        custom = status.custom_status or {}
        data_result = custom.get("data_result", {})
        story_id = custom.get("story_id", "")
        if not data_result:
            output = status.output or {}
            data_result = output.get("data_result", {})
            story_id = output.get("story_id", "")
        if not data_result:
            return func.HttpResponse(json.dumps({"error": "No data test results found"}), status_code=404, mimetype="application/json")

        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        hdr_font = Font(bold=True, size=14, color="FFFFFF")
        hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        bdr = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        ws.merge_cells("A1:B1")
        ws["A1"] = f"Data Test Report - {story_id}"
        ws["A1"].font = hdr_font
        ws["A1"].fill = hdr_fill
        ws["A3"], ws["B3"] = "Total Tests", data_result.get("total", 0)
        ws["A4"], ws["B4"] = "Passed", data_result.get("passed", 0)
        ws["B4"].fill = pass_fill
        ws["A5"], ws["B5"] = "Failed", data_result.get("failed", 0)
        if data_result.get("failed", 0) > 0:
            ws["B5"].fill = fail_fill

        ws2 = wb.create_sheet("Test Results")
        headers = ["Test Name", "Category", "Status", "Duration", "Message"]
        for ci, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = hdr_fill
        for i, tr in enumerate(data_result.get("test_results", []), 2):
            ws2.cell(row=i, column=1, value=tr.get("name", ""))
            ws2.cell(row=i, column=2, value=tr.get("category", ""))
            sc = ws2.cell(row=i, column=3, value=tr.get("status", "").upper())
            if tr.get("status") == "passed":
                sc.fill = pass_fill
            elif tr.get("status") in ("failed", "error"):
                sc.fill = fail_fill
            ws2.cell(row=i, column=4, value=f"{tr.get('duration', 0)}s")
            ws2.cell(row=i, column=5, value=tr.get("message", ""))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"data_test_report_{story_id}.xlsx"
        return func.HttpResponse(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})
    except Exception as e:
        logger.error("download-data-report error: %s", e)
        return func.HttpResponse(json.dumps({"error": str(e)}), status_code=500, mimetype="application/json")


# ============================================================
# HTTP: Test Categories CRUD
# ============================================================
@app.route(route="test-categories", methods=["GET", "POST", "PUT", "DELETE"], auth_level=func.AuthLevel.FUNCTION)
async def test_categories(req: func.HttpRequest) -> func.HttpResponse:
    try:
        if req.method == "GET":
            categories = _get_categories_from_db()
            return func.HttpResponse(json.dumps({"categories": categories}), mimetype="application/json")

        elif req.method == "POST":
            body = req.get_json()
            cat_id = body.get("id", "")
            name = body.get("name", "")
            description = body.get("description", "")
            if not cat_id:
                return func.HttpResponse('{"error": "id required"}', status_code=400, mimetype="application/json")
            _ensure_categories_table()
            conn = _get_config_db_connection()
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM test_categories WHERE id = ?", (cat_id,))
                if cursor.fetchone()[0] == 0:
                    cursor.execute(
                        "INSERT INTO test_categories (id, name, description, enabled, is_default) "
                        "VALUES (?, ?, ?, 1, 0)",
                        (cat_id, name, description),
                    )
                    conn.commit()
            finally:
                conn.close()
            categories = _get_categories_from_db()
            return func.HttpResponse(json.dumps({"categories": categories}), mimetype="application/json")

        elif req.method == "PUT":
            body = req.get_json()
            cat_id = body.get("id", "")
            if not cat_id:
                return func.HttpResponse('{"error": "id required"}', status_code=400, mimetype="application/json")
            _ensure_categories_table()
            conn = _get_config_db_connection()
            try:
                cursor = conn.cursor()
                updates = []
                params = []
                for field in ("name", "description"):
                    if field in body:
                        updates.append(f"{field} = ?")
                        params.append(body[field])
                if "enabled" in body:
                    updates.append("enabled = ?")
                    params.append(1 if body["enabled"] else 0)
                if updates:
                    params.append(cat_id)
                    cursor.execute(
                        f"UPDATE test_categories SET {', '.join(updates)} WHERE id = ?",
                        tuple(params),
                    )
                    conn.commit()
            finally:
                conn.close()
            categories = _get_categories_from_db()
            return func.HttpResponse(json.dumps({"categories": categories}), mimetype="application/json")

        elif req.method == "DELETE":
            cat_id = req.params.get("id", "")
            if not cat_id:
                return func.HttpResponse('{"error": "id required"}', status_code=400, mimetype="application/json")
            _ensure_categories_table()
            conn = _get_config_db_connection()
            try:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM test_categories WHERE id = ?", (cat_id,))
                conn.commit()
            finally:
                conn.close()
            categories = _get_categories_from_db()
            return func.HttpResponse(json.dumps({"categories": categories}), mimetype="application/json")

        return func.HttpResponse("Method not allowed", status_code=405)
    except Exception as e:
        logger.error("test-categories error: %s", e)
        return func.HttpResponse(json.dumps({"error": str(e)}), status_code=500, mimetype="application/json")


# ============================================================
# ORCHESTRATOR: Test Pipeline
# ============================================================
@app.orchestration_trigger(context_name="context")
def test_orchestrator(context: df.DurableOrchestrationContext):
    """Unified test orchestrator: Routes to UI, Data, or Both test paths.

    Flow: Read Story → Route → Plan → Review Gate → ADO Artifacts → Execute → Report
    """
    test_input = context.get_input()
    story_id = test_input.get("story_id", "")
    input_test_type = test_input.get("test_type", "")  # optional manual override

    steps = [
        {"step": 1, "name": "Read Story", "status": "completed", "detail": test_input.get("title", "")},
        {"step": 2, "name": "Route Tests", "status": "pending", "detail": ""},
        {"step": 3, "name": "Plan Tests", "status": "pending", "detail": ""},
        {"step": 4, "name": "Review Gate", "status": "pending", "detail": ""},
        {"step": 5, "name": "ADO Artifacts", "status": "pending", "detail": ""},
        {"step": 6, "name": "Generate & Execute", "status": "pending", "detail": ""},
        {"step": 7, "name": "Report Results", "status": "pending", "detail": ""},
    ]
    test_cases = []

    def _update(step_num, status, detail=""):
        for s in steps:
            if s["step"] == step_num:
                s["status"] = status
                if detail:
                    s["detail"] = detail
        context.set_custom_status({
            "story_id": story_id, "steps": steps,
            "test_cases": test_cases, "awaiting_test_review": False,
        })

    # STEP 2: Route
    if input_test_type in ("ui", "data"):
        test_type = input_test_type
        route_result = {"test_type": input_test_type, "confidence": 1.0, "data_aspects": []}
        _update(2, "completed", f"Type: {test_type} (manual override)")
    else:
        _update(2, "in_progress", "AI classifying test type...")
        route_result = yield context.call_activity("route_test_type", test_input)
        test_type = route_result.get("test_type", "data")
        _update(2, "completed", f"Type: {test_type} (confidence: {route_result.get('confidence', 0):.0%})")

    # STEP 3: Plan
    _update(3, "in_progress", f"Planning {test_type} tests...")
    ui_plan = None
    data_plan = None

    if test_type in ("ui", "both"):
        ui_plan = yield context.call_activity("plan_tests", test_input)
        scenario_count = len(ui_plan.get("test_scenarios", []))

    if test_type in ("data", "both"):
        data_plan = yield context.call_activity("plan_data_tests", {
            **test_input,
            "data_aspects": route_result.get("data_aspects", []),
            "custom_categories": test_input.get("custom_categories", []),
        })
        data_test_count = len(data_plan.get("tests", []))

    if test_type == "ui":
        _update(3, "completed", f"{scenario_count} UI test scenarios planned")
    elif test_type == "data":
        _update(3, "completed", f"{data_test_count} data validation tests planned")
    else:
        _update(3, "completed", f"UI: {scenario_count} scenarios, Data: {data_test_count} tests")

    # Build test_cases array for review
    test_cases = []
    if ui_plan:
        for sc in ui_plan.get("test_scenarios", []):
            test_cases.append({
                "id": sc.get("id", ""),
                "name": sc.get("name", sc.get("scenario", "")),
                "category": "ui",
                "priority": sc.get("priority", "medium"),
                "status": "pending",
            })
    if data_plan:
        for dt in data_plan.get("tests", []):
            test_cases.append({
                "id": dt.get("id", ""),
                "name": dt.get("name", ""),
                "category": dt.get("category", "data"),
                "priority": dt.get("priority", "medium"),
                "status": "pending",
            })

    # STEP 4: Review Gate — pause for human approval
    _update(4, "in_progress", "Awaiting test plan approval...")
    context.set_custom_status({
        "story_id": story_id, "steps": steps,
        "test_cases": test_cases, "awaiting_test_review": True,
    })

    expiry = context.current_utc_datetime + timedelta(minutes=30)
    timeout_task = context.create_timer(expiry)
    approval_task = context.wait_for_external_event("TestPlanApproved")
    winner = yield context.task_any([approval_task, timeout_task])

    if winner == timeout_task:
        # Review timed out
        _update(4, "failed", "Review timed out after 30 minutes")
        context.set_custom_status({
            "story_id": story_id, "steps": steps,
            "test_cases": test_cases, "awaiting_test_review": False,
        })
        return {"status": "timed_out", "story_id": story_id, "reason": "Review gate timed out"}

    # Approval event received — cancel timer
    timeout_task.cancel()
    approval_data = approval_task.result

    if not isinstance(approval_data, dict) or not approval_data.get("approved", False):
        reason = approval_data.get("reason", "Declined by reviewer") if isinstance(approval_data, dict) else "Declined"
        _update(4, "failed", f"Declined: {reason}")
        context.set_custom_status({
            "story_id": story_id, "steps": steps,
            "test_cases": test_cases, "awaiting_test_review": False,
        })
        return {"status": "declined", "story_id": story_id, "reason": reason}

    _update(4, "completed", "Test plan approved")
    context.set_custom_status({
        "story_id": story_id, "steps": steps,
        "test_cases": test_cases, "awaiting_test_review": False,
    })

    # STEP 5: ADO Test Artifacts
    _update(5, "in_progress", "Creating ADO Test Plan/Suite/Cases...")
    if ui_plan:
        ado_artifacts = yield context.call_activity("create_ado_test_artifacts", ui_plan)
        _update(5, "completed", f"Plan #{ado_artifacts.get('plan_id', '')}, {ado_artifacts.get('case_count', 0)} cases")
    else:
        ado_artifacts = {}
        _update(5, "completed", "Skipped (data-only)")

    # STEP 6: Generate & Execute
    exec_result = {}

    # Mark all test cases as running
    for tc in test_cases:
        tc["status"] = "running"
    context.set_custom_status({
        "story_id": story_id, "steps": steps,
        "test_cases": test_cases, "awaiting_test_review": False,
    })

    if test_type == "data":
        _update(6, "in_progress", "Executing SQL data tests...")
        data_result = yield context.call_activity("execute_data_tests", data_plan)
        exec_result = data_result

        # Update test_cases with actual results
        result_map = {r.get("id", ""): r for r in data_result.get("test_results", [])}
        for tc in test_cases:
            r = result_map.get(tc["id"])
            if r:
                tc["status"] = r.get("status", "passed")
                tc["duration"] = r.get("duration", 0)

        _update(6, "completed", f"Passed: {data_result.get('passed',0)}, Failed: {data_result.get('failed',0)}")
        context.set_custom_status({
            "story_id": story_id, "steps": steps, "test_cases": test_cases,
            "data_result": data_result, "data_test_plan": data_plan,
            "awaiting_test_review": False,
        })

    elif test_type == "ui":
        _update(6, "in_progress", "Generating Playwright scripts...")
        ui_generated = yield context.call_activity("generate_test_code", ui_plan)
        context.set_custom_status({
            "story_id": story_id, "steps": steps, "test_cases": test_cases,
            "generated_code": ui_generated, "app_url": test_input.get("app_url", ""),
            "test_plan": ui_plan or {},
            "awaiting_upload": True, "awaiting_test_review": False,
        })
        _update(6, "in_progress", "Waiting for test results upload...")
        try:
            upload_event = yield context.wait_for_external_event("TestResultsUploaded")
        except TimeoutError:
            _update(6, "failed", "Results upload timed out")
            return {"status": "timed_out", "story_id": story_id}
        exec_result = upload_event if isinstance(upload_event, dict) else {}
        p = exec_result.get("passed", 0)
        f = exec_result.get("failed", 0)

        # Update UI test_cases with results
        ui_results = exec_result.get("test_results", [])
        ui_idx = 0
        for tc in test_cases:
            if tc.get("category") == "ui":
                if ui_idx < len(ui_results):
                    r = ui_results[ui_idx]
                    tc["status"] = r.get("status", "passed")
                    tc["duration"] = r.get("duration", 0)
                else:
                    tc["status"] = "warning"
                ui_idx += 1

        _update(6, "completed", f"Passed: {p}, Failed: {f}")

    elif test_type == "both":
        _update(6, "in_progress", "Running data tests + generating UI scripts...")
        data_result = yield context.call_activity("execute_data_tests", data_plan)
        ui_generated = yield context.call_activity("generate_test_code", ui_plan)

        # Update data test_cases with results
        result_map = {r.get("id", ""): r for r in data_result.get("test_results", [])}
        for tc in test_cases:
            if tc.get("category") != "ui":
                r = result_map.get(tc["id"])
                if r:
                    tc["status"] = r.get("status", "passed")
                    tc["duration"] = r.get("duration", 0)

        context.set_custom_status({
            "story_id": story_id, "steps": steps, "test_cases": test_cases,
            "generated_code": ui_generated, "data_result": data_result,
            "app_url": test_input.get("app_url", ""), "awaiting_upload": True,
            "test_plan": ui_plan or {}, "data_test_plan": data_plan,
            "awaiting_test_review": False,
        })
        _update(6, "in_progress", f"Data: {data_result.get('passed',0)}P/{data_result.get('failed',0)}F. Waiting for UI upload...")
        try:
            upload_event = yield context.wait_for_external_event("TestResultsUploaded")
        except TimeoutError:
            upload_event = {"passed": 0, "failed": 0, "total": 0, "test_results": []}
        ui_result = upload_event if isinstance(upload_event, dict) else {}

        # Update UI test_cases with results
        ui_results = ui_result.get("test_results", [])
        ui_idx = 0
        for tc in test_cases:
            if tc.get("category") == "ui":
                if ui_idx < len(ui_results):
                    r = ui_results[ui_idx]
                    tc["status"] = r.get("status", "passed")
                    tc["duration"] = r.get("duration", 0)
                else:
                    tc["status"] = "warning"
                ui_idx += 1

        exec_result = {
            "passed": data_result.get("passed", 0) + ui_result.get("passed", 0),
            "failed": data_result.get("failed", 0) + ui_result.get("failed", 0),
            "errors": data_result.get("errors", 0) + ui_result.get("errors", 0),
            "total": data_result.get("total", 0) + ui_result.get("total", 0),
            "test_results": data_result.get("test_results", []) + ui_result.get("test_results", []),
            "data_result": data_result,
            "ui_result": ui_result,
        }
        _update(6, "completed", f"Total: {exec_result['passed']}P/{exec_result['failed']}F")

    # STEP 7: Report
    _update(7, "in_progress", "Reporting to ADO + Teams...")
    report_result = yield context.call_activity("report_test_results", {
        "exec_result": exec_result,
        "ado_artifacts": ado_artifacts,
        "test_plan": ui_plan or {},
        "story_id": story_id,
        "work_item_id": test_input.get("work_item_id"),
        "title": test_input.get("title", ""),
    })
    _update(7, "completed", f"Reported. Bugs: {len(report_result.get('bug_ids', []))}")

    return {"status": "completed", "story_id": story_id, **exec_result, "bug_ids": report_result.get("bug_ids", [])}


# ============================================================
# ACTIVITY FUNCTIONS
# ============================================================
@app.activity_trigger(input_name="payload")
def route_test_type(payload: dict) -> dict:
    config = AppConfig.from_env()
    from tester.test_router import TestRouter
    router = TestRouter(config)
    return router.classify(payload)


@app.activity_trigger(input_name="payload")
def plan_tests(payload: dict) -> dict:
    config = AppConfig.from_env()
    from tester.test_planner import TestPlannerAgent
    agent = TestPlannerAgent(config)
    return agent.plan_tests(payload, app_url=payload.get("app_url", ""))


@app.activity_trigger(input_name="payload")
def create_ado_test_artifacts(payload: dict) -> dict:
    from tester.ado_test_client import ADOTestClient
    client = ADOTestClient()
    return client.create_full_test_hierarchy(payload)


@app.activity_trigger(input_name="payload")
def generate_test_code(payload: dict) -> dict:
    config = AppConfig.from_env()
    from tester.test_generator import TestGeneratorAgent
    agent = TestGeneratorAgent(config)
    return agent.generate_tests(payload)


@app.activity_trigger(input_name="payload")
def plan_data_tests(payload: dict) -> dict:
    config = AppConfig.from_env()
    from tester.data_test_planner import DataTestPlanner
    planner = DataTestPlanner(config)
    return planner.plan_data_tests(
        story=payload, data_aspects=payload.get("data_aspects", []),
        custom_categories=payload.get("custom_categories", []),
    )


@app.activity_trigger(input_name="payload")
def execute_data_tests(payload: dict) -> dict:
    config = AppConfig.from_env()
    from tester.data_test_executor import DataTestExecutor
    executor = DataTestExecutor()
    return executor.execute_tests(payload)


@app.activity_trigger(input_name="payload")
def report_test_results(payload: dict) -> dict:
    from tester.ado_test_client import ADOTestClient
    from tester.test_reporter import report_results

    exec_result = payload.get("exec_result", {})
    ado_artifacts = payload.get("ado_artifacts", {})
    test_plan = payload.get("test_plan", {})
    story_id = payload.get("story_id", "")
    work_item_id = payload.get("work_item_id")
    title = payload.get("title", "")
    bug_ids = []
    ado_client = ADOTestClient()

    try:
        plan_id = ado_artifacts.get("plan_id")
        suite_id = ado_artifacts.get("suite_id")
        tc_map = ado_artifacts.get("test_case_map", {})
        if plan_id and suite_id:
            run = ado_client.create_test_run(plan_id, suite_id, f"Auto-Run - {story_id}")
            run_id = run["id"]
            scenarios = test_plan.get("test_scenarios", [])
            ado_results = []
            for i, tr in enumerate(exec_result.get("test_results", [])):
                tc_id_key = scenarios[i].get("id", "") if i < len(scenarios) else ""
                tc_id = tc_map.get(tc_id_key, 0)
                if tc_id:
                    ado_results.append({"test_case_id": tc_id, "status": tr.get("status", ""), "duration": tr.get("duration", 0), "message": tr.get("message", "")})
            if ado_results:
                ado_client.update_test_results(run_id, ado_results)
            ado_client.complete_test_run(run_id)
    except Exception as e:
        logger.warning("ADO test run update failed: %s", e)

    try:
        for tr in exec_result.get("test_results", []):
            if tr.get("status") in ("failed", "error"):
                repro = f"Test: {tr.get('name', '')}\nError: {tr.get('message', '')}\nStory: {story_id}"
                bug = ado_client.create_bug(title=f"[Auto] Test Failed: {tr.get('name', '')[:80]}", repro_steps=repro, story_id=work_item_id, priority=2)
                bug_ids.append(bug["id"])
    except Exception as e:
        logger.warning("Bug creation failed: %s", e)

    try:
        report_results(story_id, title, exec_result, bug_ids)
    except Exception as e:
        logger.warning("Teams notification failed: %s", e)

    return {"bug_ids": bug_ids, "status": "reported"}
