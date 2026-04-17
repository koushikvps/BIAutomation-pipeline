"""Excel Report Generator — creates a formatted test results workbook."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_excel_report(
    story_id: str,
    scenarios: list[dict],
    results: dict,
    output_path: Path,
):
    """Generate a formatted Excel report with test results."""
    wb = Workbook()

    # ─── Summary Sheet ───
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30

    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:B1")
    ws["A1"] = f"Test Report — {story_id}"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Story ID"
    ws["B3"] = story_id
    ws["A4"] = "Date"
    ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A5"] = "Total Tests"
    ws["B5"] = results.get("total", 0)
    ws["A6"] = "Passed"
    ws["B6"] = results.get("passed", 0)
    ws["B6"].fill = pass_fill
    ws["A7"] = "Failed"
    ws["B7"] = results.get("failed", 0)
    if results.get("failed", 0) > 0:
        ws["B7"].fill = fail_fill
    ws["A8"] = "Errors"
    ws["B8"] = results.get("errors", 0)
    ws["A9"] = "Duration"
    ws["B9"] = f"{results.get('elapsed_seconds', 0)}s"

    for row in range(3, 10):
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = border
        ws[f"B{row}"].border = border

    # ─── Results Sheet ───
    ws2 = wb.create_sheet("Test Results")
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 10
    ws2.column_dimensions["F"].width = 60

    headers = ["Test ID", "Test Name", "Category", "Status", "Duration", "Message"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.border = border

    for i, tr in enumerate(results.get("test_results", []), 2):
        name_parts = tr.get("name", "").split(": ", 1)
        tc_id = name_parts[0] if len(name_parts) > 1 else f"TC-{i-1:03d}"
        tc_name = name_parts[1] if len(name_parts) > 1 else name_parts[0]

        ws2.cell(row=i, column=1, value=tc_id).border = border
        ws2.cell(row=i, column=2, value=tc_name).border = border
        ws2.cell(row=i, column=3, value=tr.get("category", "")).border = border

        status_cell = ws2.cell(row=i, column=4, value=tr.get("status", "").upper())
        status_cell.border = border
        if tr.get("status") == "passed":
            status_cell.fill = pass_fill
            status_cell.font = Font(color="006100")
        elif tr.get("status") in ("failed", "error"):
            status_cell.fill = fail_fill
            status_cell.font = Font(color="9C0006")

        ws2.cell(row=i, column=5, value=f"{tr.get('duration', 0):.1f}s").border = border
        ws2.cell(row=i, column=6, value=tr.get("message", "")).border = border

    # ─── Scenarios Sheet (Test Plan) ───
    ws3 = wb.create_sheet("Test Plan")
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 50
    ws3.column_dimensions["C"].width = 15
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 8
    ws3.column_dimensions["F"].width = 20
    ws3.column_dimensions["G"].width = 20
    ws3.column_dimensions["H"].width = 40

    plan_headers = ["ID", "Title", "Category", "Priority", "Step#", "Action", "Target", "Expected"]
    for col, h in enumerate(plan_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.border = border

    row_num = 2
    for sc in scenarios:
        for step in sc.get("steps", []):
            ws3.cell(row=row_num, column=1, value=sc.get("id", "")).border = border
            ws3.cell(row=row_num, column=2, value=sc.get("title", "")).border = border
            ws3.cell(row=row_num, column=3, value=sc.get("category", "")).border = border
            ws3.cell(row=row_num, column=4, value=sc.get("priority", "")).border = border
            ws3.cell(row=row_num, column=5, value=step.get("step", "")).border = border
            ws3.cell(row=row_num, column=6, value=step.get("action", "")).border = border
            ws3.cell(row=row_num, column=7, value=step.get("target", "")).border = border
            ws3.cell(row=row_num, column=8, value=step.get("expected", "")).border = border
            row_num += 1

    wb.save(output_path)
